import openpyxl, csv, codecs
import pathlib
import random
from openpyxl import Workbook, load_workbook

def wellness_dialog_for_autoregressive(data_path = "../data/wellness_dialog_dataset.xlsx", save_path = "../data/wellness_dialog_for_autoregressive.txt"):
  wb = load_workbook(filename=data_path)
  ws = wb[wb.sheetnames[0]]
  with codecs.open(save_path, "w", "utf-8") as f:
    type_buffer = None
    q_buffer = []
    a_buffer = []

    for row in ws.iter_rows(): # row = ['category', 'question', 'answer']
      # pass the first row
      if row[0].value == '구분':
        continue

      if type_buffer == row[0].value or type_buffer is None:
        type_buffer = row[0].value
        q_buffer.append(row[1].value)
        if row[2].value is not None:
          a_buffer.append(row[2].value)
      else:
        for i in range(len(q_buffer)):
          for j in range(len(a_buffer)):
            line_to_write = q_buffer[i] + "\t" + a_buffer[j] + "\n"
            f.writelines(line_to_write)
        type_buffer = row[0].value
        q_buffer = []
        a_buffer = []
        q_buffer.append(row[1].value)
        if row[2].value is not None:
          a_buffer.append(row[2].value)

def chatbot_wellness_dialog_for_autoregressive(data_path1 = "../data/wellness_dialog_dataset.xlsx", data_path2="../data/ChatbotData.csv", save_path = "../data/chatbot_wellness_dialog_for_autoregressive.txt"):
  wb = load_workbook(filename=data_path1)
  ws = wb[wb.sheetnames[0]]
  with codecs.open(save_path, "w", "utf-8") as f:
    type_buffer = None
    q_buffer = []
    a_buffer = []

    for row in ws.iter_rows(): # row = ['category', 'question', 'answer']
      # pass the first row
      if row[0].value == '구분':
        continue

      if type_buffer == row[0].value or type_buffer is None:
        type_buffer = row[0].value
        q_buffer.append(row[1].value)
        if row[2].value is not None:
          a_buffer.append(row[2].value)
      else:
        for i in range(len(q_buffer)):
          for j in range(len(a_buffer)):
            line_to_write = q_buffer[i] + "\t" + a_buffer[j] + "\n"
            f.writelines(line_to_write)
        type_buffer = row[0].value
        q_buffer = []
        a_buffer = []
        q_buffer.append(row[1].value)
        if row[2].value is not None:
          a_buffer.append(row[2].value)

    with open(data_path2, 'r', encoding='UTF8') as csvfile:
      csv_reader = csv.reader(csvfile)
      for row in csv_reader:
        if row[0] == 'Q':
          continue
          
        line_to_write = row[0] + "\t" + row[1] + "\n"
        f.writelines(line_to_write) # save question-answer pairs

def wellness_data_for_text_classification(data_path = "../data/wellness_dialog_dataset.xlsx", save_path = "../data/wellness_data_for_text_classification.txt"):
  wb = load_workbook(filename=data_path)
  ws = wb[wb.sheetnames[0]]
  idx_buffer = ' '
  idx = -1

  with codecs.open(save_path, "w", "utf-8") as f:
    for row in ws.iter_rows():
      if row[0].value == '구분':
        continue
      else:
        if idx_buffer != row[0].value:
          idx += 1
          idx_buffer = row[0].value
        line_to_write = row[1].value + "\t" + str(idx) + "\t" + row[0].value + "\n"
        f.writelines(line_to_write)

if __name__ == "__main__":
  root_path = str(pathlib.Path(__file__).parent.absolute()) + "/../data"
  xlsx_path = root_path + "/wellness_dialog_dataset.xlsx"
  csv_path = root_path + "/ChatbotData.csv"
  well_path = root_path + "/wellness_dialog_for_autoregressive.txt"
  chat_well_path = root_path + "/chatbot_wellness_dialog_for_autoregressive.txt"
  cls_path = root_path + "/wellness_data_for_text_classification.txt"

  wellness_dialog_for_autoregressive(data_path=xlsx_path, save_path=well_path)
  chatbot_wellness_dialog_for_autoregressive(data_path1 = xlsx_path, data_path2=csv_path, save_path = chat_well_path)
  wellness_data_for_text_classification(data_path=xlsx_path, save_path=cls_path)